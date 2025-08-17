import random
from dataclasses import dataclass, field
import numpy as np  
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from collections import deque

def main(Tests, Scores, Durations,
          S_min, w_score, w_duration, M, iteration, cutpoints_num,
          mutation_rate, k, crossover_rate, live_rate
          , mating_pool_scalar, output_num):
    #scores 包含了每一个requirement的score，长度为1 * m
    #Durations包含了每个test的duration，长度为1 * n
    #M是杂交池的大小
    #每个requirement有一个score，每一个test有一个duration
    #w_score是得分权重 见fitness_vec函数
    #w_duration是时长权重 见fitness_vec函数
#S_min是得分门槛
#x是向量，w_score和w_time是权重
#M是杂交池的大小
#iteration是迭代次数
#cutpoints_num是等位基因切割点数量
#mutation_rate是变异率，见crossbreed_parents函数
#k是k-tournament参数，见select_parents函数
#crossover_rate是交叉成功率，见crossbreed_parents函数
#live_rate是父代在后代的生还率，见replace_parents函数
#mating_pool_scalar是控制杂交多少后代的参数。杂交出的后代的数量为mating_pool_scalar * M （见crossbreed_parents函数)
#output_num决定输出几个向量，见main函数
    global Tests_names, Requirements_names
    Tests_matrix = np.asarray([row.x for row in Tests], dtype = np.uint8)
    Tests_matrix_transpose = Tests_matrix.T
    n, m = Tests_matrix.shape #n是一共有几个tests，m是一共有几个requirements
    Scores = np.asarray(Scores, dtype = float)
    Durations = np.asarray(Durations, dtype = float)
    S_max = float(np.sum(Scores))
    T_max = float(np.sum(Durations))
    eps = 10 ** (-12)

    def random_vector(n):#n是test的数量
        #随机生成np阵列(二进制向量S)，用来作为初始的父母
        return np.random.randint(0, 2, size = n, dtype = np.uint8)  

    def duration_vec(x_vec: np.ndarray) -> float:
        #方案x的时长，越低越好
        return float(np.dot(x_vec, Durations)) #点乘。将每个test的时长加起来

    def repetitions_vec(x_vec: np.ndarray) -> np.ndarray:
    # 每个 requirement 被选中的 tests 覆盖了几次 -> (m,)
    # 等价于 (Tests_matrix.T @ x_vec)
        return (Tests_matrix.T @ x_vec.astype(int)).astype(int)

    def coverage_vec(x_vec: np.ndarray) -> np.ndarray:
        # 是否覆盖（>=1 次）
        return repetitions_vec(x_vec) > 0   # (m,) bool

    def score_vec(x_vec: np.ndarray) -> float:
        # 总得分 = 覆盖到的 requirement 的 Scores 之和
        cov = coverage_vec(x_vec)
        return float(np.dot(cov.astype(float), Scores))
    
    def fitness_vec(x_vec: np.ndarray) :
        #适应度函数，综合duration和score，越高越好
        s_part = (score_vec(x_vec)) / max(S_max, eps)
        t_part = (duration_vec(x_vec) - 0) / max(T_max, eps)
        return w_score * s_part - w_duration * t_part 

    #测试方案x, x包含数个tests。x是numpy阵列
    @dataclass(eq = False)
    class scheme():
        x:np.ndarray
        fitness: float = field(init=False)
        score: float = field(init=False)
        duration: float = field(init=False)
        coverage: np.ndarray = field(init=False)
        repetitions: np.ndarray = field(init=False)

        def __post_init__(self):
            self.fitness = fitness_vec(self.x)
            self.coverage = coverage_vec(self.x)
            self.score = score_vec(self.x)
            self.duration = duration_vec(self.x)
            self.repetitions = repetitions_vec(self.x)

        def __hash__(self):
            return hash(self.x.tobytes())
        
        def __eq__(self, other):
            return isinstance(other, scheme) and np.array_equal(self.x,
            other.x)
        
    def select_parents(parents, offspring = None):
        #k-tournament：每次抽取k个人，选择最优秀（fitness最高的）的。
        # 直到杂交池有M个。
        assert len(parents) >= k, "k 不能大于当前父代规模"
        mating_pool = []
        if offspring is not None:
            population = offspring
            mask = np.random.rand(len(parents)) < live_rate #判断父母是否活下来了
            population.extend(parents[mask])#如果活下来了，就把它extend到population里
        else:
            population = parents
        def select(population):
            #进行一次k-tournament
            winner = max(random.sample(population, k),
                          key = lambda ind: ind.fitness)
            return winner
        max_try = mating_pool_scalar * M                # 随便设个上限
        tries = 0
        selected = set()
        while len(mating_pool) < M and tries < max_try:
            winner = select(population)
            if winner not in selected:
                selected.add(winner)
                mating_pool.append(winner)
            tries += 1
        while len(mating_pool) < M:
            mating_pool.append(random.choice(population))
        return mating_pool
    
    def crossbreed_parents(parents):
        #杂交
        def mutation(x_vec: np.ndarray):
            #变异率是mutation_rate
            mask = np.random.rand(n) < mutation_rate #生成长度为n的数组
            #每个元素是rand()是否小于mutation_rate的布尔代数
            x_vec[mask] ^= 1 #mask中true的元素进行翻转
        
        def crossbreed(father: scheme, mother: scheme):
             #father, mother, self都是长度为n只为0或1的向量
             #有l个cut point，即l + 1个等位基因
             #只有crossover_rate的概率进行交叉操作
            cutpoints = sorted(random.sample(range(1, n), cutpoints_num)) if cutpoints_num > 0 else []
            cutpoints.append(n)
            start = 0
            child1, child2 = father.x.copy(), mother.x.copy()
            for cutpoint in cutpoints:
                if n > 1 and random.random() <= crossover_rate: #交叉成功
                    child1[start: cutpoint] = mother.x[start: cutpoint]#等位基因片段
                    child2[start: cutpoint] = father.x[start: cutpoint]
                start = cutpoint
            mutation(child1)
            mutation(child2)
            return scheme(child1), scheme(child2)
        
        offspring = []
        assert len(parents) >= 2, "父代个体数需 >= 2 才能交叉"
        while len(offspring) < mating_pool_scalar * M:
            father, mother = random.sample(parents, 2)
            child1, child2 = crossbreed(father, mother)
            offspring.append(child1)
            if len(offspring) < mating_pool_scalar * M:
                offspring.append(child2)
        return offspring
    
    def replace_parents(parents, offspring):
        #新一代杂交池
        alive_mask = np.random.rand(len(parents)) < live_rate
        survivors = [p for p, alive in zip(parents, alive_mask) if alive]

        merged = survivors + offspring
        merged.sort(key=lambda ind: ind.fitness, reverse=True)
        selected = set()
        output = []
        i = 0
        while len(output) < m:
            individual = merged[i]
            if individual not in selected:
                output.append(individual)
                selected.add(individual)
            i += 1
        return output
    #——以下是主程序——
    parents = [scheme(random_vector(n)) for v in range(M)] 
    best_score = 0
    i = 1
    
    while i <= iteration or best_score < S_min:
        print('pass' + str(i))
        mating_pool = select_parents(parents)
        offspring = crossbreed_parents(mating_pool)
        parents = replace_parents(parents, offspring)
        best_score = max(parents, key = lambda ind: ind.fitness).score
        i += 1
    bests = list(reversed(sorted(parents, key = lambda ind: ind.fitness)))[:output_num]

    #——这里是输出的地方，调excel——
    new_workbook = Workbook()
    if "Sheet" in new_workbook.sheetnames:
        del new_workbook["Sheet"]
    for i in range(output_num):
        best = bests[i]
        worksheet = new_workbook.create_sheet()
        worksheet.title = f'Scheme{i + 1}'

        #——第一部分：输出remain和remove了哪些tests
        worksheet['A1'] = ' Fitness'
        worksheet['A2'] = best.fitness
        worksheet['B1'] = ' Score'
        worksheet['B2'] = best.score
        worksheet['C1'] = ' Overall Duration'
        worksheet['C2'] = best.duration

        worksheet['A5'] = ' Remained tests'
        worksheet['B5'] = ' Durations'
        worksheet['C5'] = ' Removed tests'
        worksheet['D5'] = ' Durations'
        worksheet['E5'] = ' Covered rate'
       
        remain_index, remove_index = 6, 6
        for test_index, test in enumerate(best.x):
            if test == 1: #放在remain里
                worksheet['A' + str(remain_index)] = Tests_names[test_index]
                worksheet['B' + str(remain_index)] = Durations[test_index]
                remain_index += 1
            elif test == 0: #放在remove里
                worksheet['C' + str(remove_index)] = Tests_names[test_index]
                worksheet['D' + str(remove_index)] = Durations[test_index]
                worksheet['E' + str(remove_index)] = str(sum(best.coverage & Tests_matrix[test_index]) / sum(Tests_matrix[test_index]) * 100) + '%' 
                #从F列开始输出哪些tests能替代removed
                def bfs(target, selected_indices, dist = {}):
                     #找到selected tests (best.x)中最少的能覆盖removed test的方案
                    pass

                def array_to_mask(a: np.ndarray) -> int:
                    a = np.asarray(a, dtype = bool)
                    mask = 0
                    for i, v in enumerate(a):
                        if v: #第i位如果是1，那么mask的第i位也应该是1
                            mask |= (1 << i)
                    return mask
                
                target_mask = array_to_mask(best.coverage & Tests_matrix[test_index])
                Selected_indices = np.flatnonzero(best.x)
                masks = {i: array_to_mask(Tests_matrix[i] & Tests_matrix[test_index]) for i in Selected_indices}
                remove_index += 1
        
        #——第二部分：输出覆盖了哪些requirements，每个重复了几次，以及在哪些tests里面重复了
        #——第三部分：输出丢失了哪些requirements，以及哪些tests会覆盖这些requirements
        part2_start_row = str(4 + int(max(remain_index, remove_index)))
        part3_start_row = str(int(4 + int(part2_start_row) + best.score))
        worksheet['A' + part2_start_row] = 'Covered requirements'
        worksheet['B' + part2_start_row] = 'Repeat times'
        worksheet['A' + part3_start_row] = 'Not covered requirements'
        covered_index = int(part2_start_row) + 1
        not_covered_index = int(part3_start_row) + 1
        for i in range(m):
            requirement_name = Requirements_names[i]
            repeat_times = best.repetitions[i]
            involved_tests = Tests_matrix_transpose[i] #通过转置矩阵找出该requirement可以被哪些tests覆盖
            #大小应该是repeat_times
            if repeat_times >= 1: #说明这个requirement被best覆盖到了
                intersection = involved_tests & best.x #找出方案涉及的tests与该requirement涉及的tests之间的交集
                intersection_names = [Tests_names[v] for v in range(n) if intersection[v] == 1]#找到test的名字
                worksheet['A' + str(covered_index)] = requirement_name
                worksheet['B' + str(covered_index)] = repeat_times
                for j in range(repeat_times):
                    col_idx = get_column_letter(3 + j) #从"C"开始，一直到"C" + j结束
                    worksheet[col_idx + part2_start_row] = 'Test name'
                    worksheet[col_idx + str(covered_index)] = intersection_names[j]
                covered_index += 1
            elif repeat_times == 0:
                worksheet['A' + str(not_covered_index)] = requirement_name
                involved_tests_names = [Tests_names[v] for v in range(n) if involved_tests[v] == 1]
                for j in range(sum(involved_tests)):
                    col_idx = get_column_letter(2 + j) #从"B"开始，一直到"B" + j结束
                    worksheet[col_idx + part3_start_row] = 'Test name'
                    worksheet[col_idx + str(not_covered_index)] = involved_tests_names[j]
                not_covered_index += 1
    import os

    base_dir = os.path.dirname(os.path.abspath(__file__))  # 脚本所在目录
    out_path = os.path.join(base_dir, "GA_schemes.xlsx")

    new_workbook.save(out_path)
    print(f"保存到: {out_path}")

    return



path = input("请输入excel路径：").strip('\"')
path = path.replace('\\', '/')
workbook = load_workbook(path, data_only= True)
worksheet = workbook.active

def get_merged_value(row, col):
    for merged_cell_index in worksheet.merged_cells.ranges:
        if merged_cell_index.min_row <= row <= merged_cell_index.max_row and merged_cell_index.min_col <= col <= merged_cell_index.max_col:
            return worksheet.cell(merged_cell_index.min_row, merged_cell_index.min_col).value
    return worksheet.cell(row, col).value

@dataclass
class Test():
    x: np.array
    name: str

Tests_names, test_name_collected = [], False
Durations, durations_collected = [], False
Requirements_names, requirement_name_collected = [], False
Tests, data_collected = [], False

col_idx = 1
while get_merged_value(1, col_idx) != None:
    merged_cell_value = str(get_merged_value(1, col_idx)).lower()
    if 'Test Name'.lower() in merged_cell_value and (not test_name_collected):
        test_name_collected = True
        row_idx = 1
        while get_merged_value(row_idx, col_idx) != None:
            cell_value = worksheet[get_column_letter(col_idx) + str(row_idx)].value
            if cell_value == None:
                row_idx += 1
                continue
            elif 'test name'.lower() in str(cell_value).lower():
                row_idx += 1
                continue
            else:
                Tests_names.append(worksheet[get_column_letter(col_idx) + str(row_idx)].value)
            row_idx += 1

    elif 'Test Time'.lower() in merged_cell_value and (not durations_collected):
        durations_collected = True
        row_idx = 1
        while get_merged_value(row_idx, col_idx) != None:
            cell_value = worksheet[get_column_letter(col_idx) + str(row_idx)].value
            if cell_value == None:
                row_idx += 1
                continue
            elif 'test time'.lower() in str(cell_value).lower():
                row_idx += 1
                continue
            else:
                Durations.append(worksheet[get_column_letter(col_idx) + str(row_idx)].value)
            row_idx += 1
        Durations= np.array(Durations)
        assert len(Durations) == len(Tests_names)

    elif 'Feature'.lower() in merged_cell_value and (not requirement_name_collected):
        requirement_name_collected = True
        row_idx = 2
        cell_value = worksheet[get_column_letter(col_idx) + str(row_idx)].value
        while cell_value == None:
            row_idx += 1
        requirement_start_row_index = row_idx
        index = col_idx
        while get_merged_value(1, index) != None and'feature'.lower() in get_merged_value(1, index).lower():
            row_idx = requirement_start_row_index
            requirement_name = []
            while get_merged_value(row_idx, index) != None and get_merged_value(row_idx, index) != 0 and get_merged_value(row_idx, index) != 1:
                requirement_name.append(get_merged_value(row_idx, index))
                row_idx += 1
            index += 1
            Requirements_names.append('_'.join(requirement_name))

    if 'Feature'.lower() in merged_cell_value and (not data_collected):
        data_collected = True
        start = get_column_letter(col_idx) + str(row_idx)
        end = get_column_letter(col_idx + len(Requirements_names) - 1) + str(row_idx + len(Durations) - 1)
        for name_idx, row in enumerate(worksheet[start: end]):
            row = [0 if cell.value == None else 1 for cell in row]
            row = np.array(row)
            name = Tests_names[name_idx]
            Tests.append(Test(row, name))
        Tests_matrix = np.asarray([row.x for row in Tests], dtype = np.uint8)
    col_idx += 1
    

Scores = [1 for i in range(len(Requirements_names))]
S_min = int(len(Requirements_names) * 0.70 + 1)

run = 1
if run: 
    main(Tests = Tests, Scores = Scores, Durations = Durations,
          S_min = S_min, w_score = 1, w_duration = 1, M = 50, 
          iteration = 5, cutpoints_num = 5,
          mutation_rate = 0.05, k = 3, crossover_rate = 0.5, live_rate = 1,
           mating_pool_scalar= 5, output_num= 5)

# "C:\Users\1000405157\Desktop\Work\Genetic algorithm\Copy of SCR vs Test Coverage new.xlsx"
#A1, BH29

#output1:[(array([0, 0, 0, 0, 0, 0, 0, 1, 1, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, 0, 0, 0, 1, 0], dtype=uint8), 0.7000126335236865, 44.0, 670.0899999999999)
#output2:[(array([0, 0, 0, 0, 0, 0, 0, 1, 1, 0, 0, 0, 0, 1, 1, 0, 0, 0, 1, 1, 0, 0, 1, 0], dtype=uint8), 0.6999064129377942, 44.0, 670.7099999999999)


'''
end = input("请输入终点：")
 #——这里是读取数据并跑程序的地方——
def split_excel_address(address):
    address = address.strip().upper()
    col = ''.join([letter for letter in address if letter.isalpha()])
    row = ''.join([digit for digit in address if digit.isdigit()])
    return col, row

start_col, start_row = split_excel_address(start)
end_col, end_row = split_excel_address(end)
'''
